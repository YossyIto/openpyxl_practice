#!/usr/bin/env python3
import openpyxl


### ファイルの新規作成と保存
def step01():
    print("step1")

    # Excelファイルの新規作成
    wb = openpyxl.Workbook()

    wb.save('input_file_step01.xlsx')
    wb.close()

### セルの代入
def step02():
    print("step02")

    # Excelファイルの新規作成
    wb = openpyxl.Workbook()
    # ワークシートの取得 （デフォルトでは"Sheet"が作られる）
    ws = wb["Sheet"]

    # セルに値の入力 cell(row,col)
    # セルの位置は、縦=row ,横=col
    ws.cell(1,1).value = "test"
    ws.cell(1,2).value = "あああ"
    ws.cell(2,1).value = "いいい"
    ws.cell(row=1, column=3).value = "ううう"

    wb.save('input_file_step02.xlsx')
    wb.close()

### セルデータの取得
def step03():
    print("step03")
    # Excelファイルの読み込み
    wb = openpyxl.load_workbook('input_file_step02.xlsx')
    ws = wb["Sheet"]

    # 最大行、最大列
    print("max_row:", ws.max_row)
    print("max_col:", ws.max_column)

    for r in range(1, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            data = ws.cell(r,c).value
            print("r=", r, "c=", c, ",data=", data)

    wb.close()

### セルの代入2
def step04():
    print("step04")

    table_num = [1,2,3,4,5]

    # Excelファイルの新規作成
    wb = openpyxl.Workbook()
    # ワークシートの取得 （デフォルトでは"Sheet"が作られる）
    ws = wb["Sheet"]

    # 縦に配置
    num_row = 1
    for row_data in table_num:
        ws.cell(num_row,1).value = row_data
        num_row += 1

    # 横に配置
    num_col = 1
    for col_data in table_num:
        ws.cell(10,num_col).value = col_data
        num_col += 1

    # 保存
    wb.save('input_file_step04.xlsx')
    wb.close()

### テキストデータの取り込み
def step05():
    print("step05")

    # テキストファイルオープン(UTF-8形式)
    with open('./lifedata_man_2023.txt', encoding='utf-8') as f:
        # 1行読み込み（先頭行）
        line = f.readline()
        # データがあるだけ繰り返す
        while line:
            # 1行のデータを表示（改行部分を除く）
            print(line.rstrip("\n"))
            # 次の行を読み込み
            line = f.readline()   

### テキストデータの取り込み～エクセルに張り付け
def step06():
    print("step06")

    wb = openpyxl.Workbook()
    ws = wb["Sheet"]

    # テキストファイルオープン(UTF-8形式)
    with open('./lifedata_man_2023.txt', encoding='utf-8') as f:

        # 1行読み込み（先頭行）
        line = f.readline()
        # コピーする行番号を1にする
        num_line = 1
        # データがあるだけ繰り返す
        while line:
            # 1行のデータを表示（改行部分を除く）
            line_1 = line.rstrip("\n")
            # 半角スペースで分割
            line_2 = line_1.split(" ")
            print(line_2)
            # エクセル上にコピー
            num_col = 1
            for col_data in line_2:
                ws.cell(num_line,num_col).value = col_data
                num_col += 1

            # 次の行を読み込み
            line = f.readline() 
            # コピーする行番号を1追加
            num_line += 1
    
    # エクセルファイルの保存
    wb.save('input_file_step06.xlsx')
    wb.close()

### テキストデータの取り込み～エクセルに張り付け（データ部の属性を指定）
def step07():
    print("step07")

    wb = openpyxl.Workbook()
    ws = wb["Sheet"]

    # テキストファイルオープン(UTF-8形式)
    with open('./lifedata_woman_2023.txt', encoding='utf-8') as f:

        # 1行読み込み（先頭行）
        line = f.readline()
        # コピーする行番号を1にする
        num_line = 1
        # データがあるだけ繰り返す
        while line:
            # 1行のデータを表示（改行部分を除く）
            line_1 = line.rstrip("\n")
            # 半角スペースで分割
            line_2 = line_1.split(" ")
            print(line_2)

            # エクセル上にコピー
            # ヘッダ部(2行目まで)
            if(num_line <= 2):
                num_col = 1
                for col_data in line_2:
                    ws.cell(num_line,num_col).value = col_data
                    num_col += 1
            # 2行目以降は整形する
            else:
                age = int(line_2[0])
                rate = float(line_2[1])
                lx = int(line_2[2])
                ndx = int(line_2[3])
                nLx = int(line_2[4])
                Tx = int(line_2[5])
                ex = float(line_2[6])
                # セルへコピー
                ws.cell(num_line,1).value = age
                ws.cell(num_line,2).value = rate
                ws.cell(num_line,2).number_format = '0.00000'
                ws.cell(num_line,3).value = lx
                ws.cell(num_line,4).value = ndx
                ws.cell(num_line,5).value = nLx
                ws.cell(num_line,6).value = Tx
                ws.cell(num_line,7).value = ex
                ws.cell(num_line,7).number_format = '0.00'

            # 次の行を読み込み
            line = f.readline() 
            # コピーする行番号を1追加
            num_line += 1

    # オートフィルタ範囲の設定
    ws.auto_filter.ref = 'A2:G2'
    # ウィンドウ枠の固定(2行目まで)
    ws.freeze_panes = 'A3'
    
    # エクセルファイルの保存
    wb.save('input_file_step07.xlsx')
    wb.close()

#
# 動作部 
#
print("start")

#step01()
#step02()
#step03()
#step04()
#step05()
#step06()
step07()

print("end")

# eof